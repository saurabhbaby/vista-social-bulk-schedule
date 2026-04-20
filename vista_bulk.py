"""Reusable library for building Vista Social bulk-scheduling CSVs.

A batch config (Python file) declares:
  - where the media lives (source folders/files)
  - where the captions live (text sections or xlsx rows)
  - the schedule (list of posts with dates, types, keys)

`build.py` runs the pipeline: copy media into the repo, parse captions,
write the CSV, and optionally git-push.
"""

from __future__ import annotations

import csv
import fnmatch
import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

import openpyxl


VISTA_HEADERS = ["message", "type", "link", " time"]
DEFAULT_BASE_URL = (
    "https://raw.githubusercontent.com/saurabhbaby/vista-social-bulk-schedule/main"
)


class CaptionSource:
    def get(self, key: str) -> str:
        raise NotImplementedError


class TextSections(CaptionSource):
    """Captions from a text file with section headers like:

        CAROUSEL A1: Some title
        ... body with slides ...
        Caption
        <the caption text, possibly multi-paragraph>
        #hashtags

        ---
    """

    def __init__(self, path: str | Path, section_prefix: str = "CAROUSEL"):
        text = Path(path).read_text(encoding="utf-8")
        head = re.compile(rf"{section_prefix}\s+(\S+?):[^\n]*\n")
        parts = head.split(text)
        caps: dict[str, str] = {}
        for i in range(1, len(parts), 2):
            key, block = parts[i], parts[i + 1]
            m = re.search(
                r"\n\s*Caption\s*\n(.+?)(?=\n\s*---|\n\s*"
                + section_prefix
                + r"\s|\Z)",
                block,
                re.DOTALL,
            )
            if m:
                txt = m.group(1).strip()
                txt = re.sub(r"\n\s*═+.*$", "", txt, flags=re.DOTALL).strip()
                caps[key] = txt
        self._caps = caps

    def get(self, key: str) -> str:
        if key not in self._caps:
            raise KeyError(f"No caption for key: {key}")
        return self._caps[key]


class XlsxByFilename(CaptionSource):
    """Captions from an xlsx where one column holds filenames and another holds captions."""

    def __init__(
        self,
        path: str | Path,
        filename_col: str = "File Name",
        caption_col: str = "Caption",
    ):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        fn_idx = header.index(filename_col)
        cap_idx = header.index(caption_col)
        caps: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[fn_idx]:
                caps[row[fn_idx]] = (row[cap_idx] or "").strip()
        self._caps = caps

    def get(self, key: str) -> str:
        if key not in self._caps:
            raise KeyError(f"No caption for key: {key}")
        return self._caps[key]


def build_caption_source(spec: dict) -> CaptionSource:
    kind = spec["type"]
    if kind == "text_sections":
        return TextSections(spec["path"], section_prefix=spec.get("section_prefix", "CAROUSEL"))
    if kind == "xlsx_by_filename":
        return XlsxByFilename(
            spec["path"],
            filename_col=spec.get("filename_col", "File Name"),
            caption_col=spec.get("caption_col", "Caption"),
        )
    raise ValueError(f"Unknown caption source type: {kind}")


def copy_carousel(src_folder: str | Path, dest_folder: Path, pattern: str = "slide_*.png") -> list[str]:
    src = Path(src_folder)
    dest_folder.mkdir(parents=True, exist_ok=True)
    slides = sorted(p.name for p in src.iterdir() if fnmatch.fnmatch(p.name, pattern))
    for s in slides:
        shutil.copy(src / s, dest_folder / s)
    return slides


def copy_static(src_file: str | Path, dest_folder: Path) -> str:
    src = Path(src_file)
    dest_folder.mkdir(parents=True, exist_ok=True)
    shutil.copy(src, dest_folder / src.name)
    return src.name


def raw_url(base_url: str, *path_parts: str) -> str:
    return base_url.rstrip("/") + "/" + "/".join(p.strip("/") for p in path_parts)


@dataclass
class BuildResult:
    csv_path: Path
    rows: int
    carousel_rows: int
    static_rows: int


def build_batch(config: dict, batch_dir: Path, repo_root: Path) -> BuildResult:
    """Run the build against a loaded config. Returns paths/counts."""
    base_url = config.get("base_url", DEFAULT_BASE_URL)
    default_time = config.get("time", "9:00 am")
    media_dir = batch_dir / "media"
    media_dir.mkdir(parents=True, exist_ok=True)

    cap_carousel = build_caption_source(config["captions"]["carousels"])
    cap_static = build_caption_source(config["captions"]["statics"])

    rel_media = media_dir.relative_to(repo_root).as_posix()
    rows: list[dict] = []

    for item in config["schedule"]:
        t = item.get("time", default_time)
        dt = f"{item['date']} {t}"

        if item["type"] == "carousel":
            folder_name = item.get("folder_name") or Path(item["src_folder"]).name
            dest = media_dir / "carousels" / folder_name
            slides = copy_carousel(item["src_folder"], dest,
                                   pattern=item.get("slide_pattern", "slide_*.png"))
            caption = cap_carousel.get(item["caption_key"])
            for s in slides:
                rows.append({
                    "message": caption,
                    "type": "image",
                    "link": raw_url(base_url, rel_media, "carousels", folder_name, s),
                    " time": dt,
                })
        elif item["type"] == "static":
            fname = copy_static(item["src_file"], media_dir / "statics")
            caption = cap_static.get(item["caption_key"])
            rows.append({
                "message": caption,
                "type": "image",
                "link": raw_url(base_url, rel_media, "statics", fname),
                " time": dt,
            })
        else:
            raise ValueError(f"Unknown schedule item type: {item['type']!r}")

    csv_path = batch_dir / "schedule.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=VISTA_HEADERS)
        w.writeheader()
        w.writerows(rows)

    return BuildResult(
        csv_path=csv_path,
        rows=len(rows),
        carousel_rows=sum(1 for r in rows if "/carousels/" in r["link"]),
        static_rows=sum(1 for r in rows if "/statics/" in r["link"]),
    )


# ---------------------------------------------------------------------------
# Optional: schedule generator. Turn a compact item list into dated entries.
# ---------------------------------------------------------------------------

WEEKDAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def dates_mon_to_thu(start: str | date, count: int, post_days: Iterable[str] = ("Mon", "Tue", "Wed", "Thu")) -> list[str]:
    """Generate `count` dates starting from `start`, only on the given weekdays."""
    if isinstance(start, str):
        cur = datetime.strptime(start, "%Y-%m-%d").date()
    else:
        cur = start
    allowed = {WEEKDAY_NAMES.index(d) for d in post_days}
    out: list[str] = []
    while len(out) < count:
        if cur.weekday() in allowed:
            out.append(cur.isoformat())
        cur += timedelta(days=1)
    return out


def apply_dates(items: list[dict], start: str, post_days: Iterable[str] = ("Mon", "Tue", "Wed", "Thu")) -> list[dict]:
    """Attach dates to a list of schedule items, honoring post_days."""
    dates = dates_mon_to_thu(start, len(items), post_days)
    out = []
    for d, item in zip(dates, items):
        out.append({**item, "date": d})
    return out
