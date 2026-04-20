#!/usr/bin/env python3
"""Build Vista Social bulk-scheduling CSV for April 2026 brand content."""

import csv
import re
import openpyxl
from pathlib import Path

BASE_URL = "https://raw.githubusercontent.com/saurabhbaby/vista-social-bulk-schedule/main"
POST_TIME = "7:00 pm"

BRAND = Path("C:/Users/Saurabh/OneDrive/Desktop/Brand")
CAPTIONS_TXT = BRAND / "Carousels-20260420T094945Z-3-001/Carousels/Captions.txt"
STATIC_XLSX = BRAND / "Statics-20260420T094937Z-3-001/Statics/Static Captions.xlsx"

# Schedule: (date, type, file_identifier, caption_key)
SCHEDULE = [
    ("2026-04-21", "S", "founders_s1.png",             "founders_s1.png"),
    ("2026-04-22", "C", "C1",                          "A1"),
    ("2026-04-23", "S", "v2_s2.png",                   "v2_s2.png"),
    ("2026-04-27", "C", "C2",                          "A2"),
    ("2026-04-28", "S", "founders_s2.png",             "founders_s2.png"),
    ("2026-04-29", "C", "C3",                          "A3"),
    ("2026-04-30", "S", "founders_s4.png",             "founders_s4.png"),
    ("2026-05-04", "C", "C4",                          "A4"),
    ("2026-05-05", "S", "v2_s4.png",                   "v2_s4.png"),
    ("2026-05-06", "C", "C5",                          "A5"),
    ("2026-05-07", "S", "static_s3.png",               "static_s3.png"),
    ("2026-05-11", "C", "C6",                          "B2"),
    ("2026-05-12", "S", "founders_s3.png",             "founders_s3.png"),
    ("2026-05-13", "C", "C7",                          "B3"),
    ("2026-05-14", "S", "static_s2.png",               "static_s2.png"),
    ("2026-05-18", "C", "C8",                          "B4"),
    ("2026-05-19", "S", "v2_s5.png",                   "v2_s5.png"),
    ("2026-05-20", "C", "C9",                          "B5"),
    ("2026-05-21", "S", "static_s10.png",              "static_s10.png"),
    ("2026-05-25", "C", "C10",                         "B6"),
]

def parse_carousel_captions(txt_path):
    """Extract caption block (caption + hashtags) for each carousel A1..A5, B1..B6."""
    text = txt_path.read_text(encoding="utf-8")
    # Split on carousel headers; match "CAROUSEL A1:", "CAROUSEL B2:" etc.
    pattern = re.compile(r"CAROUSEL\s+([AB]\d+):[^\n]*\n")
    parts = pattern.split(text)
    # parts = [prefix, "A1", block1, "A2", block2, ...]
    captions = {}
    for i in range(1, len(parts), 2):
        key = parts[i]
        block = parts[i + 1]
        # Find "Caption" marker and take everything after it until "---" or next CAROUSEL
        m = re.search(r"\n\s*Caption\s*\n(.+?)(?=\n\s*---|\n\s*CAROUSEL\s|\Z)", block, re.DOTALL)
        if m:
            caption_text = m.group(1).strip()
            # Remove trailing section dividers
            caption_text = re.sub(r"\n\s*═+.*$", "", caption_text, flags=re.DOTALL).strip()
            captions[key] = caption_text
    return captions

def parse_static_captions(xlsx_path):
    """Extract full caption per filename from Static Captions.xlsx."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    fn_idx = header.index("File Name")
    cap_idx = header.index("Caption")
    captions = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[fn_idx]:
            continue
        captions[row[fn_idx]] = (row[cap_idx] or "").strip()
    return captions

def main():
    carousel_caps = parse_carousel_captions(CAPTIONS_TXT)
    static_caps = parse_static_captions(STATIC_XLSX)

    rows = []
    for date, typ, fid, cap_key in SCHEDULE:
        datetime_str = f"{date} {POST_TIME}"
        if typ == "C":
            caption = carousel_caps.get(cap_key, "")
            if not caption:
                raise SystemExit(f"Missing carousel caption for {cap_key}")
            # 5 rows per carousel
            for n in range(1, 6):
                url = f"{BASE_URL}/carousels/{fid}/slide_{n:02d}.png"
                rows.append({
                    "Message": caption,
                    "Media Type": "image",
                    "Media URL": url,
                    "Date/Time": datetime_str,
                })
        else:
            caption = static_caps.get(cap_key, "")
            if not caption:
                raise SystemExit(f"Missing static caption for {cap_key}")
            url = f"{BASE_URL}/statics/{fid}"
            rows.append({
                "Message": caption,
                "Media Type": "image",
                "Media URL": url,
                "Date/Time": datetime_str,
            })

    out = Path(__file__).parent / "schedule.csv"
    with out.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Message", "Media Type", "Media URL", "Date/Time"])
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote {len(rows)} rows to {out}")
    # Summary
    carousel_rows = sum(1 for r in rows if "/carousels/" in r["Media URL"])
    static_rows = sum(1 for r in rows if "/statics/" in r["Media URL"])
    print(f"  carousel rows: {carousel_rows} (expect 50)")
    print(f"  static rows:   {static_rows} (expect 10)")

if __name__ == "__main__":
    main()
